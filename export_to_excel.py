import os
import sys
import csv
import glob
import subprocess

# Check if openpyxl is installed; if not, install it
try:
    import openpyxl
except ImportError:
    print("Installing openpyxl for Excel generation...")
    # Use pip from the current virtualenv to install it cleanly
    pip_path = os.path.join(os.path.dirname(sys.executable), 'pip')
    if os.path.exists(pip_path):
        subprocess.check_call([pip_path, "install", "openpyxl"])
    else:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl


def convert_csv_to_excel(csv_dir, excel_dir):
    os.makedirs(excel_dir, exist_ok=True)
    
    csv_files = glob.glob(os.path.join(csv_dir, "*.csv"))
    print(f"📊 Found {len(csv_files)} CSV files in {csv_dir}")
    
    # Create a master workbook for all tables
    master_wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = master_wb.active
    master_wb.remove(default_sheet)
    
    for csv_file in sorted(csv_files):
        table_name = os.path.splitext(os.path.basename(csv_file))[0]
        excel_file_path = os.path.join(excel_dir, f"{table_name}.xlsx")
        
        print(f"🔄 Converting {table_name}.csv to Excel...")
        
        # Read CSV data
        rows = []
        with open(csv_file, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            rows = list(reader)
            
        total_rows = len(rows)
        if total_rows == 0:
            print(f"  ⚠️ Skipping empty table '{table_name}'")
            continue
        
        # 1. Create individual Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = table_name[:30] # Excel sheet name limit is 31 chars
        
        for r_idx, row in enumerate(rows, 1):
            for c_idx, val in enumerate(row, 1):
                # Try to convert to float/int if possible, for better Excel formatting
                try:
                    if val.isdigit():
                        val = int(val)
                    elif val.replace('.', '', 1).isdigit() and '.' in val:
                        val = float(val)
                except:
                    pass
                ws.cell(row=r_idx, column=c_idx, value=val)
                
        wb.save(excel_file_path)
        print(f"  ✅ Saved individual Excel: {excel_file_path} ({total_rows} rows)")
        
        # 2. Add to master combined workbook (only if it has data rows)
        if total_rows > 1: # Header + at least one data row
            sheet_title = table_name[:30]
            # Replace invalid sheet name chars
            for char in ['*', ':', '?', '/', '\\', '[', ']']:
                sheet_title = sheet_title.replace(char, '_')
                
            ws_master = master_wb.create_sheet(title=sheet_title)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, val in enumerate(row, 1):
                    try:
                        if val.isdigit():
                            val = int(val)
                        elif val.replace('.', '', 1).isdigit() and '.' in val:
                            val = float(val)
                    except:
                        pass
                    ws_master.cell(row=r_idx, column=c_idx, value=val)
                    
    # Save the combined master workbook
    master_excel_path = os.path.join(os.path.dirname(csv_dir), "camperscloset_legacy_master.xlsx")
    master_wb.save(master_excel_path)
    print(f"\n🎉 Saved Master Combined Excel Workbook: {master_excel_path}")

if __name__ == '__main__':
    convert_csv_to_excel('exports_legacy', 'exports_legacy_excel')
